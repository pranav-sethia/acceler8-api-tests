"""
Action items tests
Tests task assignments with quizzes and email notifications
"""
import requests
import pytest
import time
import random
import string
import io 
from openpyxl import load_workbook
from config import API_HOST, ORG_HEADERS as HEADERS, RAPIDAPI_KEY

ASSESS_URL      = f"{API_HOST}/backend/v1/assessment"
ASSESS_LIST_URL = f"{API_HOST}/backend/v1/assessments"
INBOXES_API = "https://inboxes-com.p.rapidapi.com"
MAILTM_API = "https://api.mail.tm"


INBOXES_API_TOKEN = RAPIDAPI_KEY

@pytest.fixture(scope="module")
def created_organisation():
    """Create a test organization for action item tests"""
    body = {
        "internal_name": "PranavAutoOrg",
        "name": "Pranav Org",
        "colour_theme": "DRIVEN_RED",
        "logo": "comet.jpg"
    }
    url = f"{API_HOST}/backend/v1/organisation"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    oid = r.json()["data"]["id"]
    yield oid
    # Cleanup: delete organization after tests
    requests.delete(f"{url}/{oid}", headers=HEADERS)

@pytest.fixture(scope="module")
def created_assessment():
    """Create a test assessment for action item tests"""
    body = {
        "capabilities":    ["TECHNICAL SKILLS 1"],
        "name":            "Test assessment for action‐items",
        "show_onboarding": False,
        "assessment_type": "EMPLOYEE"
    }

    r = requests.post(ASSESS_URL, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    aid = r.json()["data"]["id"]
    print(aid)  # Debug: print assessment ID
    yield aid
    # Cleanup: delete assessment after tests
    requests.delete(f"{ASSESS_URL}/{aid}", headers=HEADERS)


@pytest.fixture(scope="module")
def created_employee(created_assessment, mailtm_account):
    """Create a test employee with email account for notifications"""
    email = mailtm_account["address"]

    body = {
        "email": email,
        "name": "E1",
        "position": "A",
        "title": "A",
        "number_of_direct_reports": "1",
        "number_of_indirect_reports": "1",
        "join_date": "2025-06-01",
        "tenure": 1,
        "years_to_retirement": 1,
        "manager": {
            "email": f"mgr+{random.randint(1000,9999)}@example.com",
            "name": "Mgr",
            "position": "M",
            "title": "M"
        },
        "performance_rating": "1",
        "voice_of_customer_results": "1",
        "team_attrition_rate_current_year": "1",
        "team_attrition_rate_previous_year": "1",
        "voice_of_employee_results": "1"
    }

    url = f"{API_HOST}/backend/v1/assessment/{created_assessment}/employee"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 201, r.text
    emp_id = r.json()["data"]["id"]

    print(emp_id)  # Debug: print employee ID
    yield emp_id

    # Cleanup: delete employee after tests
    requests.delete(f"{API_HOST}/backend/v1/employee/{emp_id}", headers=HEADERS)


@pytest.fixture(scope="module")
def created_action_item(created_assessment, created_organisation, created_employee):
    """Create a test action item with quiz"""
    base         = f"{API_HOST}/backend/v1/assessment/{created_assessment}"
    create_url   = f"{base}/action-item"

    body = {
        "title": "A",
        "embed_items": [
            {
                "id": 1,
                "type": "TEXT",
                "content": ""
            }
        ],
        "attachments_title": "",
        "quiz_title": "",
        "send_date": "2025-06-10",
        "send_time": "12:34:46",
        "timezone": "Asia/Calcutta",
        "attachments": [],
        "quizzes": [
            {
                "id": 0,
                "type": "SUBJECTIVE",
                "question": "Question",
                "options": [],
                "submitted": False,
                "correct_options": [],
                "correct_options_count": 0
            }
        ],
        "recipient_emails": [],
        "send_to_all": True,
        "organisation_id": created_organisation
    }

    r = requests.post(create_url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    action_id = r.json()["data"]["id"]
    yield action_id
    # Cleanup: delete action item after tests
    requests.delete(f"{create_url}/{action_id}", headers=HEADERS)


def test_get_action_item_details(created_assessment, created_action_item):
    """Test retrieving action item details"""
    url = (
        f"{API_HOST}/backend/v1/assessment/"
        f"{created_assessment}/action-item/{created_action_item}"
    )
    r = requests.get(url, headers=HEADERS)
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["id"] == created_action_item
    assert data["title"] == "A"


def test_update_action_item(created_assessment, created_action_item, created_organisation):
    """Test updating action item information"""
    base       = f"{API_HOST}/backend/v1/assessment/{created_assessment}"
    update_url = f"{base}/action-item/{created_action_item}"

    body = {
        "title": "B",  # Updated title
        "embed_items": [
            {
                "id": 1,
                "type": "TEXT",
                "content": ""
            }
        ],
        "attachments_title": "",
        "quiz_title": "",
        "send_date": "2025-06-10",
        "send_time": "12:34:46",
        "timezone": "Asia/Calcutta",
        "attachments": [],
        "quizzes": [
            {
                "id": 0,
                "type": "SUBJECTIVE",
                "question": "Question",
                "options": [],
                "submitted": False,
                "correct_options": [],
                "correct_options_count": 0
            }
        ],
        "recipient_emails": [],
        "send_to_all": True,
        "organisation_id": created_organisation
    }

    r = requests.put(update_url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text

    # Verify the update worked
    r2 = requests.get(update_url, headers=HEADERS)
    assert r2.status_code == 200
    assert r2.json()["data"]["title"] == "B"


def test_list_action_items(created_assessment, created_action_item):
    """Test listing action items with pagination"""
    url = (
        f"{API_HOST}/backend/v1/assessment/"
        f"{created_assessment}/action-items"
        "?page_number=1&page_size=1"
    )
    r = requests.get(url, headers=HEADERS)
    assert r.status_code == 200
    page = r.json()
    items = page["data"]
    assert created_action_item in [i["id"] for i in items]


def test_delete_action_item(created_assessment, created_organisation):
    """Test deleting an action item"""
    base       = f"{API_HOST}/backend/v1/assessment/{created_assessment}"
    create_url = f"{base}/action-item"

    # Create a temporary action item to delete
    body = {
        "title": "To delete",
        "embed_items": [
            {
                "id": 1,
                "type": "TEXT",
                "content": ""
            }
        ],
        "attachments_title": "",
        "quiz_title": "",
        "send_date": "2025-06-10",
        "send_time": "12:34:46",
        "timezone": "Asia/Calcutta",
        "attachments": [],
        "quizzes": [],
        "recipient_emails": [],
        "send_to_all": True,
        "organisation_id": created_organisation
    }
    r0 = requests.post(create_url, json=body, headers=HEADERS)
    assert r0.status_code == 200, r0.text
    did = r0.json()["data"]["id"]

    # Delete the action item
    r1 = requests.delete(f"{create_url}/{did}", headers=HEADERS)
    assert r1.status_code in (200, 204)

    # Verify it's gone
    r2 = requests.get(f"{create_url}/{did}", headers=HEADERS)
    assert r2.status_code in (403, 404)


@pytest.fixture(scope="module")
def mailtm_account():
    """Create a temporary email account for testing notifications"""
    r = requests.get(f"{MAILTM_API}/domains")
    assert r.status_code == 200, r.text
    domains = r.json()["hydra:member"]
    assert domains, "No Mail.tm domains available"
    domain = domains[0]["domain"]

    # Generate random email and password
    local = "".join(random.choices(string.ascii_lowercase + string.digits, k=8))
    address = f"{local}@{domain}"
    password = "".join(random.choices(string.ascii_letters + string.digits, k=12))

    # Create account
    r = requests.post(
        f"{MAILTM_API}/accounts",
        json={"address": address, "password": password}
    )
    assert r.status_code == 201, r.text
    account_id = r.json()["id"]

    # Get auth token
    r = requests.post(
        f"{MAILTM_API}/token",
        json={"address": address, "password": password}
    )
    assert r.status_code == 200, r.text
    token = r.json()["token"]

    print(address)  # Debug: print email address

    yield {
        "address": address,
        "password": password,
        "token": token,
        "account_id": account_id
    }

    # Cleanup: delete email account after tests
    requests.delete(
        f"{MAILTM_API}/accounts/{account_id}",
        headers={"Authorization": f"Bearer {token}"}
    )


def test_send_reminder_and_check_inbox(
    created_assessment,
    created_action_item,
    created_employee,
    mailtm_account
):
    """Test sending action item reminder and verifying email delivery"""
    send_url = (
        f"{API_HOST}/backend/v1/assessment/"
        f"{created_assessment}/action-item/{created_action_item}/send-reminder"
    )
    r = requests.post(send_url,
                      json={"employee_ids": [created_employee]},
                      headers=HEADERS)
    assert r.status_code == 200, r.text

    # Check email inbox for the reminder
    headers = {"Authorization": f"Bearer {mailtm_account['token']}"}
    messages = []
    for _ in range(60):  # Wait up to 60 seconds for email
        r = requests.get(f"{MAILTM_API}/messages", headers=headers)
        assert r.status_code in (200, 201), r.text
        messages = r.json()["hydra:member"]
        if messages:
            break
        time.sleep(1)

    assert messages, "No messages received in Mail.tm inbox"

    # Get message details
    msg_id = messages[0]["id"]
    r = requests.get(f"{MAILTM_API}/messages/{msg_id}", headers=headers)
    assert r.status_code == 200, r.text
    msg = r.json()

    subject = msg.get("subject", "")
    body_text = msg.get("text") or "".join(msg.get("html", []))

def test_download_action_item_responses(created_assessment, created_action_item):
    """Test downloading action item responses as Excel file"""
    url = (
        f"{API_HOST}/backend/v1/assessment/"
        f"{created_assessment}/action-item/{created_action_item}/responses/export"
    )
    r = requests.get(url, headers=HEADERS)
    assert r.status_code == 200, r.text

    # Parse Excel file
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active

    assert ws.title == "B"  # Updated title from test

    # Check headers
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == ["SN", "Employee Name", "Completed", "Q1: Question"]

    # Find employee row
    found = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == "E1":
            found = row
            break

    assert found is not None, "Row for employee 'E1' not found in export"
    assert found[2] in ("Yes", "No")  # Completion status

def test_employee_can_view_and_submit_action_item(created_assessment, created_action_item, created_employee):
    """Test employee workflow: view and submit action item"""
    base_url = f"{API_HOST}/backend/v1/assessment/{created_assessment}/employee/{created_employee}"

    # List action items for employee
    list_url = f"{base_url}/action-items"
    r_list = requests.get(list_url, headers=HEADERS)
    assert r_list.status_code == 200
    action_item_ids = [item['id'] for item in r_list.json()['data']]
    assert created_action_item in action_item_ids

    # Get action item details
    details_url = f"{base_url}/action-item/{created_action_item}"
    r_details = requests.get(details_url, headers=HEADERS)
    assert r_details.status_code == 200
    assert r_details.json()['data']['id'] == created_action_item

    # Submit response
    submit_body = {
        "response": [
            {
                "quiz_id": "0",
                "quiz_type": "SUBJECTIVE",
                "answer": "This is my answer from the automated test.",
                "options": []
            }
        ],
        "completed": True
    }
    
    submit_url = f"{details_url}/submit"
    r_submit = requests.post(submit_url, json=submit_body, headers=HEADERS)
    
    assert r_submit.status_code == 200, r_submit.text


@pytest.fixture(scope="module")
def action_item_with_varied_statuses(created_assessment, created_organisation):
    """Create action item with employees in different completion statuses"""
    # Create the Action Item
    create_url = f"{ASSESS_URL}/{created_assessment}/action-item"
    action_item_body = {
        "title": "Status Test Action Item",
        "embed_items": [{"id": 1, "type": "TEXT", "content": ""}],
        "send_to_all": True,
        "send_date": "2025-01-01",
        "send_time": "10:00:00",
        "timezone": "Asia/Calcutta",
        "organisation_id": created_organisation,
        "quizzes": [{"id": "0", "type": "SUBJECTIVE", "question": "Status Q?"}]
    }
    r_ai = requests.post(create_url, json=action_item_body, headers=HEADERS)
    assert r_ai.status_code == 200, f"Failed to create action item: {r_ai.text}"
    action_item_id = r_ai.json()["data"]["id"]

    # Create two employees with different statuses
    employees = {}
    for status in ["DONE", "NOT_DONE"]:
        rand_suffix = "".join(random.choices(string.ascii_lowercase, k=4))
        emp_body = {
            "email": f"ai_status_{status.lower()}_{rand_suffix}@example.com",
            "name": f"Employee {status}",
            "manager": {"email": f"ai_mgr_{rand_suffix}@example.com", "name": "AI Manager"}
        }
        r_emp = requests.post(f"{ASSESS_URL}/{created_assessment}/employee", json=emp_body, headers=HEADERS)
        assert r_emp.status_code == 201, f"Failed to create employee: {r_emp.text}"
        employees[status] = {"id": r_emp.json()["data"]["id"], "name": emp_body["name"]}
    
    # Submit response for one employee only
    done_employee_id = employees["DONE"]["id"]
    submit_url = f"{ASSESS_URL}/{created_assessment}/employee/{done_employee_id}/action-item/{action_item_id}/submit"
    submit_body = {
        "response": [{"quiz_id": "0", "answer": "Done"}],
        "completed": True
    }
    r_submit = requests.post(submit_url, json=submit_body, headers=HEADERS)
    assert r_submit.status_code == 200, f"Failed to submit response: {r_submit.text}"

    time.sleep(1)  # Wait for status updates

    yield {"assessment_id": created_assessment, "action_item_id": action_item_id, "employees": employees}


def test_admin_can_see_action_item_status(action_item_with_varied_statuses):
    """Test that admin can see correct completion status for action items"""
    assessment_id = action_item_with_varied_statuses["assessment_id"]
    action_item_id = action_item_with_varied_statuses["action_item_id"]
    employees = action_item_with_varied_statuses["employees"]
    done_employee_id = employees["DONE"]["id"]
    not_done_employee_id = employees["NOT_DONE"]["id"]
    
    details_url = f"{ASSESS_URL}/{assessment_id}/action-item/{action_item_id}"
    r_details = requests.get(details_url, headers=HEADERS)
    assert r_details.status_code == 200
    
    action_item_data = r_details.json()["data"]
    
    # Map employee IDs to completion status
    status_map = {
        recipient["id"]: recipient.get("is_completed", False) 
        for recipient in action_item_data.get("recipients", [])
    }
    
    # Verify completion statuses
    assert status_map.get(done_employee_id) is True
    assert status_map.get(not_done_employee_id) is False