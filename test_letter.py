import requests
import pytest
import time
import random
import string
import io
from openpyxl import load_workbook
from config import API_HOST, ORG_HEADERS as HEADERS, RAPIDAPI_KEY

LETTER_TEMPLATE_URL = f"{API_HOST}/backend/v1/assessment"
MAILTM_API = "https://api.mail.tm"


@pytest.fixture(scope="module")
def created_organisation():
    """Create an organization for testing."""
    body = {
        "internal_name": "LetterTestOrg",
        "name": "Letter Test Organization",
        "colour_theme": "DRIVEN_RED",
        "logo": "comet.jpg"
    }
    url = f"{API_HOST}/backend/v1/organisation"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    oid = r.json()["data"]["id"]
    yield oid
    requests.delete(f"{url}/{oid}", headers=HEADERS)


@pytest.fixture(scope="module")
def created_assessment():
    """Create an assessment for testing letters."""
    body = {
        "capabilities": ["TECHNICAL SKILLS 1"],
        "name": "Test assessment for letters",
        "show_onboarding": False,
        "assessment_type": "EMPLOYEE"
    }
    url = f"{API_HOST}/backend/v1/assessment"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    aid = r.json()["data"]["id"]
    yield aid
    requests.delete(f"{url}/{aid}", headers=HEADERS)


@pytest.fixture(scope="module")
def mailtm_account():
    """Create a temporary email account for testing."""
    r = requests.get(f"{MAILTM_API}/domains")
    assert r.status_code == 200, r.text
    domains = r.json()["hydra:member"]
    assert domains, "No Mail.tm domains available"
    domain = domains[0]["domain"]
    
    local = "".join(random.choices(string.ascii_lowercase + string.digits, k=8))
    address = f"{local}@{domain}"
    password = "".join(random.choices(string.ascii_letters + string.digits, k=12))
    
    r = requests.post(
        f"{MAILTM_API}/accounts",
        json={"address": address, "password": password}
    )
    assert r.status_code == 201, r.text
    account_id = r.json()["id"]
    
    r = requests.post(
        f"{MAILTM_API}/token",
        json={"address": address, "password": password}
    )
    assert r.status_code == 200, r.text
    token = r.json()["token"]
    
    
    yield {
        "address": address,
        "password": password,
        "token": token,
        "account_id": account_id
    }
    
    requests.delete(
        f"{MAILTM_API}/accounts/{account_id}",
        headers={"Authorization": f"Bearer {token}"}
    )


@pytest.fixture(scope="module")
def created_employee(created_assessment, mailtm_account):
    """Create an employee for testing."""
    email = mailtm_account["address"]
    body = {
        "email": email,
        "name": "Letter Test Employee",
        "position": "Developer",
        "title": "Senior Developer",
        "number_of_direct_reports": "2",
        "number_of_indirect_reports": "5",
        "join_date": "2025-01-01",
        "tenure": 2,
        "years_to_retirement": 30,
        "manager": {
            "email": f"manager+{random.randint(1000,9999)}@example.com",
            "name": "Test Manager",
            "position": "Engineering Manager",
            "title": "Engineering Manager"
        },
        "performance_rating": "4",
        "voice_of_customer_results": "85",
        "team_attrition_rate_current_year": "10",
        "team_attrition_rate_previous_year": "12",
        "voice_of_employee_results": "90"
    }
    
    url = f"{API_HOST}/backend/v1/assessment/{created_assessment}/employee"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 201, r.text
    emp_id = r.json()["data"]["id"]
    
    yield emp_id
    
    requests.delete(f"{API_HOST}/backend/v1/employee/{emp_id}", headers=HEADERS)


@pytest.fixture(scope="module")
def created_letter_template(created_assessment):
    """Create a letter template for testing."""
    body = {
        "number_of_days": 10,
        "reason_for_joining": "abcd",
        "users_wish": "pqrs"
    }
    
    url = f"{LETTER_TEMPLATE_URL}/{created_assessment}/letter/template"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    
    yield created_assessment
    

def test_create_letter_template(created_assessment):
    """Test creating a letter template."""
    body = {
        "number_of_days": 15,
        "reason_for_joining": "abcd",
        "users_wish": "pqrs"
    }
    
    url = f"{LETTER_TEMPLATE_URL}/{created_assessment}/letter/template"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    
    data = r.json()["data"]
    assert data["number_of_day"] == 15 
    assert data["reason_for_joining"] == "abcd"
    assert data["users_wish"] == "pqrs"


def test_update_letter_template(created_letter_template):
    """Test updating a letter template."""
    body = {
        "number_of_days": 20,
        "reason_for_joining": "Updated reason for joining",
        "users_wish": "Updated wishes for the future"
    }
    
    url = f"{LETTER_TEMPLATE_URL}/{created_letter_template}/letter/template"
    
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    
    data = r.json()["data"]
    assert data.get("number_of_day") == 20


def test_list_letters(created_assessment, created_letter_template):
    """Test listing letters with pagination."""
    url = (
        f"{LETTER_TEMPLATE_URL}/{created_assessment}/letters"
        "?page_number=1&page_size=1"
    )
    
    r = requests.get(url, headers=HEADERS)
    assert r.status_code == 200, r.text
    
    page = r.json()
    assert "data" in page
    
    letters = page["data"]
    assert isinstance(letters, list)


def test_send_letter_notification(created_assessment, created_employee, created_letter_template):
    """Test sending letter notification to employees."""
    body = {
        "employee_ids": [created_employee]
    }
    
    url = f"{LETTER_TEMPLATE_URL}/{created_assessment}/letter/notify"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    
    data = r.json()
    assert "data" in data or "message" in data


def test_send_letter_notification_and_check_inbox(
    created_assessment, 
    created_employee, 
    created_letter_template,
    mailtm_account
):
    """Test sending letter notification and verify email receipt."""
    body = {
        "employee_ids": [created_employee]
    }
    
    url = f"{LETTER_TEMPLATE_URL}/{created_assessment}/letter/notify"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    
    headers = {"Authorization": f"Bearer {mailtm_account['token']}"}
    messages = []
    
    for _ in range(60):
        r = requests.get(f"{MAILTM_API}/messages", headers=headers)
        assert r.status_code in (200, 201), r.text
        messages = r.json()["hydra:member"]
        if messages:
            break
        time.sleep(1)
    
    assert messages, "No messages received in Mail.tm inbox"
    
    msg_id = messages[0]["id"]
    r = requests.get(f"{MAILTM_API}/messages/{msg_id}", headers=headers)
    assert r.status_code == 200, r.text
    
    msg = r.json()
    subject = msg.get("subject", "")
    body_text = msg.get("text") or "".join(msg.get("html", []))
    
    assert subject == "Complete Your Letter to Yourself – Your Journey Starts Here"
    assert body_text[:200] == """Hi Letter Test Employee,

Your capability journey is about to begin, but there's one important step
left—your letter to yourself. This letter is your personal commitment, a message
from your present s"""
    
def test_download_letter_excel(created_assessment, created_letter_template):
    """Test downloading letter responses as Excel."""
    url = f"{LETTER_TEMPLATE_URL}/{created_assessment}/letter/export"
    
    r = requests.get(url, headers=HEADERS)
    assert r.status_code == 200, r.text
    
    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active
    
    assert ws is not None, "No active worksheet found"
    
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    assert len(headers) > 0, "Excel should have at least one column"
    assert headers == ['Employee Name', 'Employee Email', 'Submission Date', 'In the next [input number] days, the one thing that I hope to achieve', 'start', 'stop', 'continue', 'I want to show up as [input reason to join the program] who', 'The moments in my work or life that I want to [input what user wish to be] are', 'sign']
    
def test_letter_template_with_special_characters(created_assessment):
    """Test creating letter template with special characters."""
    body = {
        "number_of_days": 30,
        "reason_for_joining": "Growth & Development! @Company #2025",
        "users_wish": "Best of luck! May you achieve 100% success & happiness."
    }
    
    url = f"{LETTER_TEMPLATE_URL}/{created_assessment}/letter/template"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    
    data = r.json()["data"]
    assert data.get("number_of_day") == 30


def test_letter_notification_multiple_employees(created_assessment, created_letter_template):
    """Test sending letter notification to multiple employees."""
    employees = []
    
    for i in range(2):
        email = f"test{random.randint(1000,9999)}@example.com"
        body = {
            "email": email,
            "name": f"Test Employee {i+1}",
            "position": "Developer",
            "title": "Developer",
            "number_of_direct_reports": "0",
            "number_of_indirect_reports": "0",
            "join_date": "2025-01-01",
            "tenure": 1,
            "years_to_retirement": 35,
            "manager": {
                "email": f"mgr{random.randint(1000,9999)}@example.com",
                "name": "Manager",
                "position": "Manager",
                "title": "Manager"
            },
            "performance_rating": "3",
            "voice_of_customer_results": "75",
            "team_attrition_rate_current_year": "5",
            "team_attrition_rate_previous_year": "5",
            "voice_of_employee_results": "80"
        }
        
        url = f"{API_HOST}/backend/v1/assessment/{created_assessment}/employee"
        r = requests.post(url, json=body, headers=HEADERS)
        assert r.status_code == 201, r.text
        emp_id = r.json()["data"]["id"]
        employees.append(emp_id)
    
    body = {
        "employee_ids": employees
    }
    
    url = f"{LETTER_TEMPLATE_URL}/{created_assessment}/letter/notify"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
        
    for emp_id in employees:
        requests.delete(f"{API_HOST}/backend/v1/employee/{emp_id}", headers=HEADERS)


def test_letter_pagination(created_assessment, created_letter_template):
    """Test letter listing with different pagination parameters."""
    for page_size in [1, 5, 10]:
        url = (
            f"{LETTER_TEMPLATE_URL}/{created_assessment}/letters"
            f"?page_number=1&page_size={page_size}"
        )
        
        r = requests.get(url, headers=HEADERS)
        assert r.status_code == 200, r.text
        
        page = r.json()
        assert "data" in page
        letters = page["data"]
        
        assert len(letters) <= page_size, f"Expected max {page_size} items, got {len(letters)}"
