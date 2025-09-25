"""
Prepare content tests
Tests pre-assessment learning materials and scheduling
"""
import requests
import pytest
import io
import time
import random
import string
from openpyxl import load_workbook
from datetime import datetime, timedelta, timezone
from config import API_HOST, ORG_HEADERS as HEADERS, RAPIDAPI_KEY

ASSESS_URL = f"{API_HOST}/backend/v1/assessment"
INBOXES_API = "https://inboxes-com.p.rapidapi.com"
MAILTM_API = "https://api.mail.tm"


INBOXES_API_TOKEN = RAPIDAPI_KEY

@pytest.fixture(scope="module")
def created_organisation():
    """Create an organization for prepare content tests"""
    body = {
        "internal_name": f"PrepContentOrg{random.randint(1000,9999)}",
        "name": "Prepare Content Org",
        "colour_theme": "DRIVEN_RED"
    }
    url = f"{API_HOST}/backend/v1/organisation"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    oid = r.json()["data"]["id"]
    yield oid
    # Cleanup: delete organization after tests
    requests.delete(f"{url}/{oid}", headers=HEADERS)

@pytest.fixture(scope="module")
def created_assessment():
    """Create an assessment for prepare content tests"""
    body = {
        "capabilities":    ["TECHNICAL SKILLS 1"],
        "name":            "Test assessment for prepare-content",
        "show_onboarding": False,
        "assessment_type": "EMPLOYEE"
    }
    r = requests.post(ASSESS_URL, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    aid = r.json()["data"]["id"]
    yield aid
    # Cleanup: delete assessment after tests
    requests.delete(f"{ASSESS_URL}/{aid}", headers=HEADERS)


@pytest.fixture(scope="module")
def created_employee(created_assessment):
    """Create an employee for prepare content tests"""
    url = f"{API_HOST}/backend/v1/assessment/{created_assessment}/employee"

    body = {
        "email":                     "test_pranav_email",
        "name":                      "E1",
        "position":                  "A",
        "title":                     "A",
        "number_of_direct_reports":  "1",
        "number_of_indirect_reports":"1",
        "join_date":                 "2025-06-01",
        "tenure":                    1,
        "years_to_retirement":       1,
        "manager": {
            "email":                      "test_pranav_manager_email",
            "name":                       "A",
            "position":                   "A",
            "title":                      "A",
            "number_of_direct_reports":   None,
            "number_of_indirect_reports": None,
            "join_date":                  None,
            "tenure":                     None,
            "years_to_retirement":        None,
            "performance_rating":         None,
            "voice_of_customer_results":  None,
            "team_attrition_rate_current_year":   None,
            "team_attrition_rate_previous_year":  None,
            "voice_of_employee_results":          None
        },
        "performance_rating":           "1",
        "voice_of_customer_results":    "1",
        "team_attrition_rate_current_year":  "1",
        "team_attrition_rate_previous_year": "1",
        "voice_of_employee_results":    "1"
    }

    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 201, r.text
    eid = r.json()["data"]["id"]
    yield eid
    # Cleanup: delete employee after tests
    requests.delete(f"{url}/{eid}", headers=HEADERS)


@pytest.fixture(scope="module")
def created_prepare_content(created_assessment, created_employee):
    """Create a prepare content item for testing"""
    base       = f"{API_HOST}/backend/v1/assessment/{created_assessment}"
    create_url = f"{base}/prepare-content"

    body = {
        "title": "A",
        "embed_items": [
            {
                "id": 1,
                "type": "TEXT",
                "content": "A"
            }
        ],
        "attachments_title": "",
        "quiz_title": "",
        "send_date": "2025-06-11",
        "send_time": "12:54:10",
        "timezone": "Asia/Calcutta",
        "attachments": [],
        "quizzes": [
            {
                "id": 0,
                "type": "SUBJECTIVE",
                "question": "A",
                "options": [],
                "submitted": False,
                "correct_options": [],
                "correct_options_count": 0
            }
        ],
        "recipient_emails": [],
        "send_to_all": True,
        "organisation_id": "4540c249-64bc-4356-b6fc-37ac600d3627"
    }

    r = requests.post(create_url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    pc_id = r.json()["data"]["id"]
    yield pc_id

    # Cleanup: delete prepare content after tests
    requests.delete(f"{create_url}/{pc_id}", headers=HEADERS)


def test_list_prepare_contents(created_assessment, created_prepare_content):
    """Test listing prepare content items"""
    url = f"{API_HOST}/backend/v1/assessment/{created_assessment}/prepare-contents"
    r = requests.get(url, headers=HEADERS)
    assert r.status_code == 200
    items = r.json()
    ids = [i["id"] for i in items["data"]]
    assert created_prepare_content in ids


def test_get_prepare_content_details(created_assessment, created_prepare_content):
    """Test retrieving prepare content details"""
    url = (
        f"{API_HOST}/backend/v1/assessment/"
        f"{created_assessment}/prepare-content/{created_prepare_content}"
    )
    r = requests.get(url, headers=HEADERS)
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["id"] == created_prepare_content
    assert data["title"] == "A"


def test_update_prepare_content(created_assessment, created_prepare_content):
    """Test updating prepare content"""
    base       = f"{API_HOST}/backend/v1/assessment/{created_assessment}"
    update_url = f"{base}/prepare-content/{created_prepare_content}"

    body = {
        "title": "B",  # Updated title
        "embed_items": [
            {
                "id": 1,
                "type": "TEXT",
                "content": "A"
            }
        ],
        "attachments_title": "",
        "quiz_title": "",
        "send_date": "2025-06-11",
        "send_time": "12:54:10",
        "timezone": "Asia/Calcutta",
        "attachments": [],
        "quizzes": [
            {
                "id": 0,
                "type": "SUBJECTIVE",
                "question": "A",
                "options": [],
                "submitted": False,
                "correct_options": [],
                "correct_options_count": 0
            }
        ],
        "recipient_emails": [],
        "send_to_all": True,
        "organisation_id": "4540c249-64bc-4356-b6fc-37ac600d3627"
    }

    r = requests.put(update_url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text

    # Verify the update worked
    r2 = requests.get(update_url, headers=HEADERS)
    assert r2.status_code == 200
    assert r2.json()["data"]["title"] == "B"


def test_delete_prepare_content(created_assessment):
    """Test deleting prepare content"""
    base       = f"{API_HOST}/backend/v1/assessment/{created_assessment}"
    create_url = f"{base}/prepare-content"

    # Create a temporary prepare content item
    body = {
        "title": "To delete",
        "embed_items": [
            {
                "id": 1,
                "type": "TEXT",
                "content": "X"
            }
        ],
        "attachments_title": "",
        "quiz_title": "",
        "send_date": "2025-06-11",
        "send_time": "12:54:10",
        "timezone": "Asia/Calcutta",
        "attachments": [],
        "quizzes": [],
        "recipient_emails": [],
        "send_to_all": True,
        "organisation_id": "4540c249-64bc-4356-b6fc-37ac600d3627"
    }

    r0 = requests.post(create_url, json=body, headers=HEADERS)
    assert r0.status_code == 200, r0.text
    pid = r0.json()["data"]["id"]

    # Delete the item
    r1 = requests.delete(f"{create_url}/{pid}", headers=HEADERS)
    assert r1.status_code in (200, 204)

    # Verify it's gone
    r2 = requests.get(f"{create_url}/{pid}", headers=HEADERS)
    assert r2.status_code in (403, 404)

@pytest.fixture(scope="module")
def mailtm_account():
    r = requests.get(f"{MAILTM_API}/domains")
    assert r.status_code == 200, r.text
    domains = r.json()["hydra:member"]
    assert domains, "No Mail.tm domains available"
    domain = domains[0]["domain"]

    local = "".join(random.choices(string.ascii_lowercase + string.digits, k=8))
    address = f"{local}@{domain}"
    password = "".join(random.choices(string.ascii_letters + string.digits, k=12))

    r = requests.post(
        f"{MAILTM_API}/accounts",
        json={"address": address, "password": password}
    )
    assert r.status_code == 201, r.text
    account_id = r.json()["id"]

    r = requests.post(
        f"{MAILTM_API}/token",
        json={"address": address, "password": password}
    )
    assert r.status_code == 200, r.text
    token = r.json()["token"]

    print(address)

    yield {
        "address": address,
        "password": password,
        "token": token,
        "account_id": account_id
    }

    requests.delete(
        f"{MAILTM_API}/accounts/{account_id}",
        headers={"Authorization": f"Bearer {token}"}
    )


def test_download_action_item_responses(created_assessment, created_prepare_content):
    url = (
        f"{API_HOST}/backend/v1/assessment/"
        f"{created_assessment}/prepare-content/{created_prepare_content}/responses/export"
    )
    r = requests.get(url, headers=HEADERS)
    assert r.status_code == 200, r.text

    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == ['SN', 'Employee Name', 'Completed', 'Q1: A']

    found = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == "E1":
            found = row
            break

    assert found == (1.0, 'E1', 'No', '')

def test_employee_can_view_and_submit_prepare_content(created_assessment, created_prepare_content, created_employee):
    base_url = f"{ASSESS_URL}/{created_assessment}/employee/{created_employee}"

    list_url = f"{base_url}/prepare-contents"
    r_list = requests.get(list_url, headers=HEADERS)
    assert r_list.status_code == 200
    content_ids = [item['id'] for item in r_list.json()['data']]
    assert created_prepare_content in content_ids

    details_url = f"{base_url}/prepare-content/{created_prepare_content}"
    r_details = requests.get(details_url, headers=HEADERS)
    assert r_details.status_code == 200
    assert r_details.json()['data']['id'] == created_prepare_content

    submit_url = f"{details_url}/submit"
    submit_body = {
        "response": [{
            "quiz_id": "0",
            "quiz_type": "SUBJECTIVE",
            "answer": "Employee response to prepare content quiz.",
            "options": []
        }],
        "completed": True
    }
    r_submit = requests.post(submit_url, json=submit_body, headers=HEADERS)
    assert r_submit.status_code == 200, r_submit.text



def test_send_reminder_and_check_inbox(created_assessment, created_prepare_content, mailtm_account):
    create_emp_url = f"{ASSESS_URL}/{created_assessment}/employee"
    employee_email = mailtm_account["address"]

    emp_body = {
        "email": employee_email,
        "name": "Reminder Recipient",
        "position": "Test Position",
        "title": "Test Title",
        "join_date": "2025-01-01",
        "manager": {
            "email": f"reminder.manager@example.com",
            "name": "Reminder Manager"
        }
    }
    
    r_emp = requests.post(create_emp_url, json=emp_body, headers=HEADERS)
    assert r_emp.status_code == 201, r_emp.text # This should now pass
    employee_id = r_emp.json()["data"]["id"]

    reminder_url = f"{ASSESS_URL}/{created_assessment}/prepare-content/{created_prepare_content}/send-reminder"
    reminder_body = { "employee_ids": [employee_id] }
    r_remind = requests.post(reminder_url, json=reminder_body, headers=HEADERS)
    assert r_remind.status_code == 200

    headers = {"Authorization": f"Bearer {mailtm_account['token']}"}
    messages = []
    for _ in range(60): 
        r_mail = requests.get(f"{MAILTM_API}/messages", headers=headers)
        assert r_mail.status_code in (200, 201)
        messages = r_mail.json()["hydra:member"]
        if messages:
            break
        time.sleep(1)

    assert messages, "Reminder email was not received in the inbox"


@pytest.fixture(scope="module")
def prepare_content_with_varied_statuses(created_assessment):
    assessment_id = created_assessment
    
    create_url = f"{ASSESS_URL}/{assessment_id}/prepare-content"
    pc_body = {
        "title": "Status Test Content", "send_to_all": True, "send_date": "2025-01-01", "send_time": "12:00",
        "organisation_id": HEADERS.get("organisation_id"),
        "quizzes": [{"id": 0, "type": "SUBJECTIVE", "question": "Status Q?"}]
    }
    r_pc = requests.post(create_url, json=pc_body, headers=HEADERS)
    assert r_pc.status_code == 200
    prepare_content_id = r_pc.json()["data"]["id"]

    employees = {}
    for status in ["NOT_DONE", "DONE"]:
        rand_suffix = "".join(random.choices(string.ascii_lowercase, k=4))
        emp_body = {
            "email": f"pc_status_{status.lower()}_{rand_suffix}@example.com", "name": f"Employee {status}",
            "manager": {"email": f"pc_mgr_{rand_suffix}@example.com", "name": "PC Manager"}
        }
        r_emp = requests.post(f"{ASSESS_URL}/{assessment_id}/employee", json=emp_body, headers=HEADERS)
        assert r_emp.status_code == 201
        employees[status] = {"id": r_emp.json()["data"]["id"], "name": emp_body["name"]}

    done_employee_id = employees["DONE"]["id"]
    submit_url = f"{ASSESS_URL}/{assessment_id}/employee/{done_employee_id}/prepare-content/{prepare_content_id}/submit"
    submit_body = {"response": [{"quiz_id": "0", "answer": "Done"}], "completed": 'true'}
    r_submit = requests.post(submit_url, json=submit_body, headers=HEADERS)
    assert r_submit.status_code == 200

    publish_url = f"{ASSESS_URL}/{assessment_id}/result/visibility"
    publish_body = {"result_visibility": "MANAGER_AND_EMPLOYEE"}
    r_publish = requests.put(publish_url, json=publish_body, headers=HEADERS)
    assert r_publish.status_code == 200, f"Failed to publish result: {r_publish.text}"

    verify_url = f"{ASSESS_URL}/{assessment_id}/employee/{done_employee_id}/verify"
    r_verify = requests.post(verify_url, headers=HEADERS)
    assert r_verify.status_code == 200, f"Failed to verify employee: {r_verify.text}"

    time.sleep(2) 

    yield {"assessment_id": assessment_id, "employees": employees}


def test_admin_can_see_prepare_content_status(prepare_content_with_varied_statuses):
    """
    Tests that the admin can see the correct completion status by checking
    the 'uncompleted_recipients' list in the prepare-content details.
    """
    assessment_id = prepare_content_with_varied_statuses["assessment_id"]
    employees = prepare_content_with_varied_statuses["employees"]
    done_employee_id = employees["DONE"]["id"]
    not_done_employee_id = employees["NOT_DONE"]["id"]
    
    list_url = f"{ASSESS_URL}/{assessment_id}/prepare-contents"
    r_list = requests.get(list_url, headers=HEADERS)
    assert r_list.status_code == 200, r_list.text
    
    prepare_content_data = None
    for item in r_list.json()["data"]:
        if item.get("title") == "Status Test Content":
            prepare_content_data = item
            break
            
    assert prepare_content_data is not None, "Could not find the test prepare-content item."
    assert prepare_content_data.get("completed_count") == 1
    
    uncompleted_ids = [
        recipient["id"] for recipient in prepare_content_data.get("uncompleted_recipients", [])
    ]
    
    assert done_employee_id not in uncompleted_ids
    assert not_done_employee_id in uncompleted_ids


def test_send_prepare_content_reminder_and_check_inbox(created_assessment, created_prepare_content, mailtm_account):
    """
    Tests the end-to-end flow of sending a reminder and verifying
    that the email is received in an inbox.
    """
    employee_email = mailtm_account["address"]
    emp_body = {
        "email": employee_email,
        "name": "Reminder Recipient",
        "manager": {"email": "reminder.mgr@example.com", "name": "Reminder Manager"}
    }
    r_emp = requests.post(f"{ASSESS_URL}/{created_assessment}/employee", json=emp_body, headers=HEADERS)
    assert r_emp.status_code == 201
    employee_id = r_emp.json()["data"]["id"]

    reminder_url = f"{ASSESS_URL}/{created_assessment}/prepare-content/{created_prepare_content}/send-reminder"
    reminder_body = {"employee_ids": [employee_id]}
    r_remind = requests.post(reminder_url, json=reminder_body, headers=HEADERS)
    assert r_remind.status_code == 200

    headers = {"Authorization": f"Bearer {mailtm_account['token']}"}
    messages = []
    for _ in range(60): 
        r_mail = requests.get(f"https://api.mail.tm/messages", headers=headers)
        assert r_mail.status_code in (200, 201)
        messages = r_mail.json()["hydra:member"]
        if messages:
            break
        time.sleep(1)

    assert messages, "Reminder email was not received in the inbox"


def test_new_employee_is_assigned_to_existing_content(created_assessment, created_organisation):
    """
    Covers: Check after add employee, will show in existing prepare content
    """
    create_url = f"{ASSESS_URL}/{created_assessment}/prepare-content"
    pc_body = {
        "title": "Existing Content for New Hires",
        "embed_items": [{"id": 1, "type": "TEXT", "content": ""}],
        "send_to_all": True,
        "send_date": "2024-01-01",
        "send_time": "12:00:00",
        "timezone": "Asia/Calcutta",
        "organisation_id": created_organisation
    }
    r_pc = requests.post(create_url, json=pc_body, headers=HEADERS)
    assert r_pc.status_code == 200, f"Failed to create prepare content: {r_pc.text}"
    prepare_content_id = r_pc.json()["data"]["id"]

    rand_suffix = "".join(random.choices(string.ascii_lowercase, k=4))
    emp_body = {
        "email": f"new_hire_{rand_suffix}@example.com", "name": "New Hire",
        "manager": {"email": f"new_hire_mgr_{rand_suffix}@example.com", "name": "Hiring Manager"}
    }
    r_emp = requests.post(f"{ASSESS_URL}/{created_assessment}/employee", json=emp_body, headers=HEADERS)
    assert r_emp.status_code == 201, f"Failed to create employee: {r_emp.text}"
    employee_id = r_emp.json()["data"]["id"]

    list_url = f"{ASSESS_URL}/{created_assessment}/employee/{employee_id}/prepare-contents"
    r_list = requests.get(list_url, headers=HEADERS)
    assert r_list.status_code == 200
    content_ids = [item['id'] for item in r_list.json()['data']]
    assert prepare_content_id in content_ids


def test_prepare_content_scheduling(created_assessment, created_organisation, created_employee):
    """
    Covers: Prepare Content is available after the Schedule AND
            Prepare content is hidden Before the Schedule
    """
    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).strftime('%Y-%m-%d')
    past_content_body = {
        "title": "Past Content", "send_to_all": True, "send_date": yesterday,
        "organisation_id": created_organisation,
        "embed_items": [{"id": 1, "type": "TEXT", "content": ""}],
        "send_time": "12:00:00", "timezone": "Asia/Calcutta"
    }
    r_past = requests.post(f"{ASSESS_URL}/{created_assessment}/prepare-content", json=past_content_body, headers=HEADERS)
    assert r_past.status_code == 200, f"Failed to create past content: {r_past.text}"
    past_content_id = r_past.json()["data"]["id"]

    tomorrow = (datetime.now(timezone.utc) + timedelta(days=1)).strftime('%Y-%m-%d')
    future_content_body = {
        "title": "Future Content", "send_to_all": True, "send_date": tomorrow,
        "organisation_id": created_organisation,
        "embed_items": [{"id": 1, "type": "TEXT", "content": ""}],
        "send_time": "12:00:00", "timezone": "Asia/Calcutta"
    }
    r_future = requests.post(f"{ASSESS_URL}/{created_assessment}/prepare-content", json=future_content_body, headers=HEADERS)
    assert r_future.status_code == 200, f"Failed to create future content: {r_future.text}"
    future_content_id = r_future.json()["data"]["id"]
    
    list_url = f"{ASSESS_URL}/{created_assessment}/employee/{created_employee}/prepare-contents"
    r_list = requests.get(list_url, headers=HEADERS)
    assert r_list.status_code == 200
    visible_content_ids = [item['id'] for item in r_list.json()['data']]

    assert past_content_id in visible_content_ids
    assert future_content_id not in visible_content_ids