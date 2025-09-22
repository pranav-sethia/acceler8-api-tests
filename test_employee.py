import requests
import pytest
import random
import string
import io
import os
import time
import zipfile
from config import API_HOST, ORG_HEADERS as HEADERS
from openpyxl import load_workbook
from openpyxl import Workbook


BASE_URL = f"{API_HOST}/backend/v1"

@pytest.fixture(scope="module")
def created_organisation():
    body = {
        "internal_name": f"TestAutoOrg{''.join(random.choices(string.digits, k=4))}",
        "name": "Test Employee Org",
        "colour_theme": "DRIVEN_RED",
        "logo": "comet.jpg"
    }
    url = f"{BASE_URL}/organisation"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    oid = r.json()["data"]["id"]
    
    yield oid
    
    requests.delete(f"{url}/{oid}", headers=HEADERS)

@pytest.fixture(scope="module")
def created_assessment(created_organisation):
    body = {
        "organisation_id": created_organisation,
        "capabilities": ["TECHNICAL SKILLS 1"],
        "name": "Employee Test Assessment",
        "show_onboarding": False,
        "assessment_type": "EMPLOYEE"
    }
    url = f"{BASE_URL}/assessment"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 200, r.text
    aid = r.json()["data"]["id"]
    
    yield aid
    
    requests.delete(f"{url}/{aid}", headers=HEADERS)

@pytest.fixture(scope="module")
def created_employee(created_assessment):
    rand_suffix = "".join(random.choices(string.ascii_lowercase + string.digits, k=6))
    employee_email = f"emp+{rand_suffix}@example.com"
    manager_email = f"mgr+{rand_suffix}@example.com"
    
    body = {
        "email": employee_email,
        "name": "Test Employee E1",
        "position": "Analyst",
        "manager": {
            "email": manager_email,
            "name": "Test Manager M1"
        }
    }
    
    url = f"{BASE_URL}/assessment/{created_assessment}/employee"
    r = requests.post(url, json=body, headers=HEADERS)
    assert r.status_code == 201, r.text
    
    data = r.json()["data"]
    employee_id = data["id"]
    employee_code = data["employee_code"]
    manager_id = data["manager"]["id"]
    
    yield {
        "employee_id": employee_id,
        "employee_code": employee_code,
        "manager_id": manager_id
    }
    
    requests.delete(f"{BASE_URL}/employee/{employee_id}", headers=HEADERS)

def test_upload_employee_assessment_file(created_assessment):
    file_path = "copy.xlsx" 

    if not os.path.exists(file_path):
        pytest.fail(f"The required test file was not found at: {file_path}")

    with open(file_path, "rb") as excel_file:
        files = {
            'file': (os.path.basename(file_path), excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        upload_headers = {
            "Authorization": HEADERS["Authorization"]
        }
        
        url = f"{BASE_URL}/assessment/{created_assessment}/employee/upload"
        r = requests.post(url, headers=upload_headers, files=files)
    
    assert r.status_code == 200, r.text
    assert r.json()["message"] == "success"

def test_get_assessment_employees_pagination(created_assessment, created_employee):
    url = f"{BASE_URL}/assessment/{created_assessment}/employees?page=1&size=10"
    r = requests.get(url, headers=HEADERS)
    
    assert r.status_code == 200, r.text
    page = r.json()
    assert page["status"] == 200
    assert page["current_page"] == 1
    
    employee_ids = [emp["id"] for emp in page["data"]]
    assert created_employee["employee_id"] in employee_ids

def test_update_assessment_employee(created_assessment, created_employee):
    employee_id = created_employee["employee_id"]
    url = f"{BASE_URL}/assessment/{created_assessment}/employee/{employee_id}"
    
    update_body = {
        "name": "Updated Test Employee",
        "position": "Senior Analyst",
    }
    
    r_put = requests.put(url, json=update_body, headers=HEADERS)
    assert r_put.status_code == 200, r_put.text
    
    r_get = requests.get(f"{BASE_URL}/employee/{employee_id}", headers=HEADERS)
    assert r_get.status_code == 200, r_get.text
    
    updated_data = r_get.json()["data"]
    assert updated_data["name"] == "Updated Test Employee"
    assert updated_data["position"] == "Senior Analyst"

def test_get_employee_by_code(created_assessment, created_employee):
    employee_code = created_employee["employee_code"]
    url = f"{BASE_URL}/assessment/{created_assessment}/employee/by_code?code={employee_code}"
    r = requests.get(url, headers=HEADERS)
    
    assert r.status_code == 200, r.text
    data = r.json()["data"]
    assert data["id"] == created_employee["employee_id"]
    assert data["employee_code"] == employee_code

def test_get_employee_by_id(created_employee):
    employee_id = created_employee["employee_id"]
    url = f"{BASE_URL}/employee/{employee_id}"
    r = requests.get(url, headers=HEADERS)
    
    assert r.status_code == 200, r.text
    assert r.json()["data"]["id"] == employee_id

def test_verify_employee_assessment(created_assessment, created_employee):
    employee_id = created_employee["employee_id"]
    url = f"{BASE_URL}/assessment/{created_assessment}/employee/{employee_id}/verify"
    
    r = requests.post(url, headers={})
    
    assert r.status_code == 200, r.text
    assert r.json()["message"] == "success"

def test_get_employee_assessments(created_employee, created_assessment):
    employee_id = created_employee["employee_id"]
    url = f"{BASE_URL}/employee/{employee_id}/assessments"
    r = requests.get(url, headers={})
    
    assert r.status_code == 200, r.text
    page = r.json()
    
    assessment_ids = [assess["id"] for assess in page["data"]]
    assert created_assessment in assessment_ids

def test_get_manager_subordinates(created_assessment, created_employee):
    manager_id = created_employee["manager_id"]
    url = f"{BASE_URL}/assessment/{created_assessment}/manager/{manager_id}/subordinates"
    r = requests.get(url, headers={})
    
    assert r.status_code == 200, r.text
    data = r.json()["data"]
    
    assert data["id"] == manager_id
    
    subordinate_ids = [sub["id"] for sub in data["subordinates"]]
    assert created_employee["employee_id"] in subordinate_ids


def test_export_employee_data(created_assessment):
    url = f"{BASE_URL}/assessment/{created_assessment}/employee/export"
    
    export_body = {
        "select_all": True
    }
    
    r = requests.post(url, json=export_body, headers=HEADERS)
    assert r.status_code == 200, r.text
    
    assert 'attachment' in r.headers.get('Content-Disposition', '')
    assert len(r.content) > 0


def test_notify_employee(created_assessment, created_employee):
    url = f"{BASE_URL}/assessment/{created_assessment}/employee/share-link"
    
    notify_body = {
        "employee_ids": [created_employee["employee_id"]]
    }
    
    r = requests.post(url, json=notify_body, headers=HEADERS)
    assert r.status_code == 200, r.text
    assert r.json()["message"] == "success"


import time

@pytest.fixture(scope="module")
def assessment_with_varied_statuses(created_assessment):
    assessment_id = created_assessment
    employee_ids = {}

    for status in ["CREATED", "COMPLETED"]:
        rand_suffix = "".join(random.choices(string.ascii_lowercase, k=4))
        body = {
            "email": f"status_{status.lower()}_{rand_suffix}@example.com", "name": f"Employee {status}",
            "manager": {"email": f"status_mgr_{rand_suffix}@example.com", "name": "Status Manager"}
        }
        url = f"{BASE_URL}/assessment/{assessment_id}/employee"
        r = requests.post(url, json=body, headers=HEADERS)
        employee_ids[status] = r.json()["data"]["id"]
    
    q_url = f"{BASE_URL}/assessment/{assessment_id}/question"
    q_body = {
        "capabilities": "TECHNICAL SKILLS 1", "capabilities_label": "TECHNICAL SKILLS 1",
        "sub_capability": "Status Test", "visibility": ["EMPLOYEE"], "type": "OBJECTIVE_SINGLE",
        "question": "What is the status?", "options": [{"value": "A", "input_required": False}],
        "ranking": [1], "sl_no": 1, "manager_question": None, "manager_question_type": None
    }
    r_q = requests.post(q_url, json=q_body, headers=HEADERS)
    assert r_q.status_code == 200, f"Failed to create question: {r_q.text}"
    question_id = r_q.json()["id"]

    submit_url = f"{API_HOST}/backend/v1/employee/{employee_ids['COMPLETED']}/assessment/{assessment_id}/response"
    submit_body = {
        "responses": [{"question_id": question_id, "options": [{"option": "A"}]}],
        "status": "SUBMITTED", "assessment_type": "EMPLOYEE"
    }
    r_submit = requests.post(submit_url, json=submit_body, headers=HEADERS)
    assert r_submit.status_code == 200, f"Failed to submit response: {r_submit.text}"
    
    publish_url = f"{BASE_URL}/assessment/{assessment_id}/result/visibility"
    publish_body = {"result_visibility": "MANAGER_AND_EMPLOYEE"}
    r_publish = requests.put(publish_url, json=publish_body, headers=HEADERS)
    assert r_publish.status_code == 200, f"Failed to publish result: {r_publish.text}"

    verify_url = f"{BASE_URL}/assessment/{assessment_id}/employee/{employee_ids['COMPLETED']}/verify"
    r_verify = requests.post(verify_url, headers=HEADERS) 
    assert r_verify.status_code == 200, f"Failed to verify employee: {r_verify.text}"

    time.sleep(2) 
    
    yield {"assessment_id": assessment_id, "employee_ids": employee_ids}

def test_admin_can_see_employee_status(assessment_with_varied_statuses):
    assessment_id = assessment_with_varied_statuses["assessment_id"]
    
    url = f"{BASE_URL}/assessment/{assessment_id}/employees?page=1&size=10"
    r = requests.get(url, headers=HEADERS)
    assert r.status_code == 200, r.text
    
    employees_data = r.json()["data"]
    
    statuses = {emp['id']: emp['status'] for emp in employees_data}
    
    assert statuses.get(assessment_with_varied_statuses["employee_ids"]["CREATED"]) == "CREATED"
    
    assert statuses.get(assessment_with_varied_statuses["employee_ids"]["COMPLETED"]) == "PROFILE_VERIFIED"

def test_notify_multiple_employees(created_assessment, created_employee):
    rand_suffix = "".join(random.choices(string.ascii_lowercase, k=4))
    body = {
        "email": f"multi_notify_{rand_suffix}@example.com",
        "name": "Multi Notify Employee",
        "manager": {"email": f"multi_mgr_{rand_suffix}@example.com", "name": "Multi Manager"}
    }
    url = f"{BASE_URL}/assessment/{created_assessment}/employee"
    r = requests.post(url, json=body, headers=HEADERS)
    second_employee_id = r.json()["data"]["id"]
    url = f"{BASE_URL}/assessment/{created_assessment}/employee/share-link"
    notify_body = {"employee_ids": [created_employee["employee_id"], second_employee_id]}
    r_notify = requests.post(url, json=notify_body, headers=HEADERS)
    assert r_notify.status_code == 200, r_notify.text
    assert r_notify.json()["message"] == "success"

def test_admin_can_copy_assessment_link(created_assessment, created_employee):
    assessment_id = created_assessment
    employee_id = created_employee["employee_id"]
    url = f"{BASE_URL}/assessment/{assessment_id}/employee/{employee_id}/result/link"
    r = requests.get(url, headers=HEADERS)
    assert r.status_code == 200, r.text
    data = r.json()["data"]
    assert "link" in data
    assert data["link"].startswith("http")


def test_update_employee_via_excel_upload(created_assessment, created_employee):
    """
    Tests updating an existing employee's data by uploading an Excel file
    that perfectly matches the required template structure and data validation.
    """
    employee_id = created_employee["employee_id"]
    
    emp_details_url = f"{BASE_URL}/employee/{employee_id}"
    r_emp = requests.get(emp_details_url, headers=HEADERS)
    assert r_emp.status_code == 200
    employee_data = r_emp.json()["data"]
    employee_email = employee_data["email"]
    manager_email = employee_data["manager"]["email"] 

    wb = Workbook()
    ws = wb.active

    ws.merge_cells('H1:I1')
    ws.merge_cells('M1:R1')
    ws.merge_cells('S1:T1')
    ws.merge_cells('V1:W1')

    header_row_1 = [
        "SN", "Name", "Email", "Country Code", "Phone Number", "Position", "Title",
        "Span of Control", None, "Join Date", "Tenure", "Years to Retirement",
        "Direct Reporting Manager", None, None, None, None, None,
        "Performance", None, "Voice of", "Team Attrition", None, "Voice of"
    ]
    ws.append(header_row_1)

    header_row_2 = [
        None, None, None, None, None, None, None,
        "Number of Direct Reports", "Number of Indirect Reports", None, None, None,
        "Name", "Position", "Title", "Email", "Country Code", "Phone Number",
        "2022", "2023", None, "2023", "2024", None
    ]
    ws.append(header_row_2)

    update_data_row = [
        1, "Updated Name", employee_email, None, None, None, None, None, None,
        None, None, None, 
        "Test Manager M1", 
        None, None, 
        manager_email, 
        None, None, None, None, None, None, None, None
    ]
    ws.append(update_data_row)
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    files = {'file': ('update.xlsx', file_stream, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    upload_headers = {"Authorization": HEADERS["Authorization"]}
    upload_url = f"{BASE_URL}/assessment/{created_assessment}/employee/upload"
    r_upload = requests.post(upload_url, headers=upload_headers, files=files)
    assert r_upload.status_code == 200, f"Upload failed: {r_upload.text}"

    r_verify = requests.get(emp_details_url, headers=HEADERS)
    assert r_verify.status_code == 200
    updated_name = r_verify.json()["data"]["name"]
    assert updated_name == "Updated Name"